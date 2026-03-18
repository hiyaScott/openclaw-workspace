#!/usr/bin/env python3
"""
编钟模拟器设计案 - 完整版 Excel 导出
包含图片说明、示意图、详细设计内容
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_complete_design_doc():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 定义样式
    title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    subtitle_font = Font(name='微软雅黑', size=11, bold=True, color='1F4E79')
    normal_font = Font(name='微软雅黑', size=10)
    note_font = Font(name='微软雅黑', size=9, color='666666', italic=True)
    code_font = Font(name='Courier New', size=9)
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    light_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    def style_header(ws, title):
        ws.append([title])
        cell = ws.cell(row=1, column=1)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.row_dimensions[1].height = 35
        ws.append([])
    
    # ===== Sheet 1: 项目概述 =====
    ws1 = wb.create_sheet("1.项目概述")
    style_header(ws1, "📋 项目概述")
    
    ws1.append(["核心定位"])
    ws1.merge_cells('A3:J3')
    ws1['A3'].font = subtitle_font
    ws1.append([])
    ws1.append(["编钟模拟器不是传统意义上的「游戏」，而是一个文化体验型音乐应用。核心价值在于："])
    ws1.merge_cells('A5:J5')
    ws1.append([])
    ws1.append(["维度", "说明"])
    ws1.merge_cells('B7:J7')
    ws1['A7'].font = Font(bold=True)
    ws1['B7'].font = Font(bold=True)
    ws1.append(["文化传承", "让现代人以互动方式接触和理解编钟这一古老乐器"])
    ws1.merge_cells('B8:J8')
    ws1.append(["音乐创作", "提供录制和分享功能，鼓励用户创作原创音乐"])
    ws1.merge_cells('B9:J9')
    ws1.append(["教育科普", "展示一钟双音的物理原理和音乐理论知识"])
    ws1.merge_cells('B10:J10')
    ws1.append([])
    
    ws1.append(["灵感来源"])
    ws1.merge_cells('A12:J12')
    ws1['A12'].font = subtitle_font
    ws1.append([])
    ws1.append(["设计参考对象：曾侯乙编钟（战国早期，1978年湖北随县出土，约公元前433年）"])
    ws1.merge_cells('A14:J14')
    ws1.append([])
    ws1.append(["参考维度", "实现方式"])
    ws1.merge_cells('B16:J16')
    ws1['A16'].font = Font(bold=True)
    ws1['B16'].font = Font(bold=True)
    ws1.append(["音阶结构", "采用五声音阶 (C-D-E-G-A) 对应宫商角徵羽"])
    ws1.merge_cells('B17:J17')
    ws1.append(["三排布局", "上层小编钟（高音）、中层中编钟（中音）、下层大编钟（低音）"])
    ws1.merge_cells('B18:J18')
    ws1.append(["一钟双音", "每口钟可发出正鼓音和侧鼓音两个不同音高"])
    ws1.merge_cells('B19:J19')
    ws1.append(["铜器音色", "使用 Web Audio API 合成带有金属谐波的钟声"])
    ws1.merge_cells('B20:J20')
    ws1.append([])
    
    ws1.append(["设计目标"])
    ws1.merge_cells('A22:J22')
    ws1['A22'].font = subtitle_font
    ws1.append([])
    ws1.append(["序号", "目标内容"])
    ws1.merge_cells('B24:J24')
    ws1['A24'].font = Font(bold=True)
    ws1['B24'].font = Font(bold=True)
    ws1.append(["1", "在移动端浏览器实现接近原生 App 的触控体验"])
    ws1.merge_cells('B25:J25')
    ws1.append(["2", "准确还原编钟「一钟双音」的声学特性"])
    ws1.merge_cells('B26:J26')
    ws1.append(["3", "提供完整的录制-回放-分享创作闭环"])
    ws1.merge_cells('B27:J27')
    ws1.append(["4", "保持界面简洁优雅，突出编钟本身的美学"])
    ws1.merge_cells('B28:J28')
    
    # ===== Sheet 2: 核心机制 =====
    ws2 = wb.create_sheet("2.核心机制")
    style_header(ws2, "⚙️ 核心机制")
    
    ws2.append(["一钟双音系统（核心特色）"])
    ws2.merge_cells('A3:J3')
    ws2['A3'].font = subtitle_font
    ws2.append([])
    ws2.append(["说明：每口编钟的两个音高相差大三度（约4个半音），仅15口钟就能覆盖30个不同音高"])
    ws2.merge_cells('A5:J5')
    ws2['A5'].font = note_font
    ws2.append([])
    
    ws2.append(["【示意图】一钟双音对比"])
    ws2.merge_cells('A7:J7')
    ws2['A7'].font = Font(bold=True, color='C00000')
    ws2.append([])
    ws2.append(["正敲（正鼓音）", "", "侧敲（侧鼓音）"])
    ws2.merge_cells('A9:B9')
    ws2.merge_cells('C9:J9')
    ws2['A9'].font = Font(bold=True, color='0070C0')
    ws2['C9'].font = Font(bold=True, color='C00000')
    ws2.append(["敲击位置", "钟体正中", "敲击位置", "钟体两侧"])
    ws2.merge_cells('B10:J10')
    ws2.append(["音高", "C4 (261.63 Hz) - 低音", "音高", "E4 (329.63 Hz) - 高音"])
    ws2.merge_cells('B11:J11')
    ws2.append(["音色特点", "浑厚、低沉", "音色特点", "清亮、通透"])
    ws2.merge_cells('B12:J12')
    ws2.append(["波形示意", "低频正弦波", "波形示意", "高频正弦波（粉色）"])
    ws2.merge_cells('B13:J13')
    ws2.append([])
    ws2.append(["音程关系：+3°（大三度，约4个半音）"])
    ws2.merge_cells('A15:J15')
    ws2['A15'].font = Font(bold=True, size=11, color='F5AF19')
    ws2.append([])
    
    ws2.append(["音区架构"])
    ws2.merge_cells('A17:J17')
    ws2['A17'].font = subtitle_font
    ws2.append([])
    ws2.append(["【示意图】三排编钟布局"])
    ws2.merge_cells('A19:J19')
    ws2['A19'].font = Font(bold=True, color='C00000')
    ws2.append([])
    ws2.append(["排层", "钟大小", "音区范围", "数量", "钟编号", "视觉设计"])
    ws2.merge_cells('F21:J21')
    for col in ['A','B','C','D','E','F']:
        ws2[f'{col}21'].font = Font(bold=True)
    ws2.append(["上层", "小编钟", "C5-A5 (高音)", "5口", "U1-U5", "尺寸最小(60px)，位置最高，颜色#FFD700"])
    ws2.merge_cells('F22:J22')
    ws2.append(["中层", "中编钟", "C4-A4 (中音)", "5口", "M1-M5", "中等尺寸(90px)，中间位置，颜色#DAA520"])
    ws2.merge_cells('F23:J23')
    ws2.append(["下层", "大编钟", "C3-A3 (低音)", "5口", "L1-L5", "尺寸最大(130px)，位置最低，颜色#B8860B"])
    ws2.merge_cells('F24:J24')
    ws2.append([])
    ws2.append(["总计：15口编钟 × 2音/钟 = 30个可用音高"])
    ws2.merge_cells('A26:J26')
    ws2['A26'].font = Font(bold=True, size=11)
    ws2.append([])
    
    ws2.append(["音色合成原理"])
    ws2.merge_cells('A28:J28')
    ws2['A28'].font = subtitle_font
    ws2.append([])
    ws2.append(["层次", "波形类型", "频率", "音量", "衰减时间", "作用"])
    ws2.merge_cells('G30:J30')
    for col in ['A','B','C','D','E','F']:
        ws2[f'{col}30'].font = Font(bold=True)
    ws2.append(["基频层", "正弦波", "f（基频）", "0.6", "3秒", "主体音色"])
    ws2.merge_cells('G31:J31')
    ws2.append(["二次谐波", "正弦波", "2f（八度）", "0.3", "2.4秒", "丰富度"])
    ws2.merge_cells('G32:J32')
    ws2.append(["三次谐波", "三角波", "3f", "0.15", "1.5秒", "金属质感"])
    ws2.merge_cells('G33:J33')
    ws2.append([])
    ws2.append(["包络设计：Attack 20ms（快速起音）→ Decay 3000ms（长衰减，模拟铜钟持续振动）"])
    ws2.merge_cells('A35:J35')
    ws2['A35'].font = note_font
    
    # ===== Sheet 3: 音频配置 =====
    ws3 = wb.create_sheet("3.音频配置表")
    style_header(ws3, "🎵 音频配置表")
    
    ws3.append(["【音频路由图】"])
    ws3.merge_cells('A3:J3')
    ws3['A3'].font = Font(bold=True, color='C00000')
    ws3.append([])
    ws3.append(["振荡器层(基频+谐波) → 增益节点(ADSR包络) → 混响节点(ConvolverNode) → 主增益(Volume:0.8) → 音频输出"])
    ws3.merge_cells('A5:J5')
    ws3['A5'].font = code_font
    ws3.append([])
    
    ws3.append(["上层（高音 C5-A5）"])
    ws3.merge_cells('A7:J7')
    ws3['A7'].font = subtitle_font
    ws3.append([])
    ws3.append(["音名", "正敲频率(Hz)", "侧敲频率(Hz)", "音分差", "正敲音名", "侧敲音名", "波形类型"])
    ws3.merge_cells('H9:J9')
    for col in ['A','B','C','D','E','F','G']:
        ws3[f'{col}9'].font = Font(bold=True)
    ws3.append(["C5", "523.25", "659.25", "400", "C5", "E5", "Sine + Triangle"])
    ws3.merge_cells('H10:J10')
    ws3.append(["D5", "587.33", "698.46", "300", "D5", "F5", "Sine + Triangle"])
    ws3.merge_cells('H11:J11')
    ws3.append(["E5", "659.25", "783.99", "300", "E5", "G5", "Sine + Triangle"])
    ws3.merge_cells('H12:J12')
    ws3.append(["G5", "783.99", "987.77", "400", "G5", "B5", "Sine + Triangle"])
    ws3.merge_cells('H13:J13')
    ws3.append(["A5", "880.00", "1046.50", "300", "A5", "C6", "Sine + Triangle"])
    ws3.merge_cells('H14:J14')
    ws3.append([])
    
    ws3.append(["中层（中音 C4-A4）"])
    ws3.merge_cells('A16:J16')
    ws3['A16'].font = subtitle_font
    ws3.append([])
    ws3.append(["音名", "正敲频率(Hz)", "侧敲频率(Hz)", "音分差", "正敲音名", "侧敲音名", "波形类型"])
    ws3.merge_cells('H18:J18')
    for col in ['A','B','C','D','E','F','G']:
        ws3[f'{col}18'].font = Font(bold=True)
    ws3.append(["C4", "261.63", "329.63", "400", "C4", "E4", "Sine + Triangle"])
    ws3.merge_cells('H19:J19')
    ws3.append(["D4", "293.66", "349.23", "300", "D4", "F4", "Sine + Triangle"])
    ws3.merge_cells('H20:J20')
    ws3.append(["E4", "329.63", "392.00", "300", "E4", "G4", "Sine + Triangle"])
    ws3.merge_cells('H21:J21')
    ws3.append(["G4", "392.00", "493.88", "400", "G4", "B4", "Sine + Triangle"])
    ws3.merge_cells('H22:J22')
    ws3.append(["A4", "440.00", "523.25", "300", "A4", "C5", "Sine + Triangle"])
    ws3.merge_cells('H23:J23')
    ws3.append([])
    
    ws3.append(["下层（低音 C3-A3）"])
    ws3.merge_cells('A25:J25')
    ws3['A25'].font = subtitle_font
    ws3.append([])
    ws3.append(["音名", "正敲频率(Hz)", "侧敲频率(Hz)", "音分差", "正敲音名", "侧敲音名", "波形类型"])
    ws3.merge_cells('H27:J27')
    for col in ['A','B','C','D','E','F','G']:
        ws3[f'{col}27'].font = Font(bold=True)
    ws3.append(["C3", "130.81", "164.81", "400", "C3", "E3", "Sine + Triangle"])
    ws3.merge_cells('H28:J28')
    ws3.append(["D3", "146.83", "174.61", "300", "D3", "F3", "Sine + Triangle"])
    ws3.merge_cells('H29:J29')
    ws3.append(["E3", "164.81", "196.00", "300", "E3", "G3", "Sine + Triangle"])
    ws3.merge_cells('H30:J30')
    ws3.append(["G3", "196.00", "246.94", "400", "G3", "B3", "Sine + Triangle"])
    ws3.merge_cells('H31:J31')
    ws3.append(["A3", "220.00", "261.63", "300", "A3", "C4", "Sine + Triangle"])
    ws3.merge_cells('H32:J32')
    
    # ===== Sheet 4: UI/UE设计 =====
    ws4 = wb.create_sheet("4.UI-UE设计")
    style_header(ws4, "🎨 UI/UE 设计")
    
    ws4.append(["设计原则"])
    ws4.merge_cells('A3:J3')
    ws4['A3'].font = subtitle_font
    ws4.append([])
    ws4.append(["原则", "说明"])
    ws4.merge_cells('B5:J5')
    ws4['A5'].font = Font(bold=True)
    ws4['B5'].font = Font(bold=True)
    ws4.append(["沉浸优先", "深色背景减少视觉干扰，突出金色编钟"])
    ws4.merge_cells('B6:J6')
    ws4.append(["横屏体验", "强制横屏模式，模拟真实编钟的宽展布局"])
    ws4.merge_cells('B7:J7')
    ws4.append(["直观操作", "无需教程，点击/触控即可发声"])
    ws4.merge_cells('B8:J8')
    ws4.append(["即时反馈", "每次敲击都有视觉和听觉双重反馈"])
    ws4.merge_cells('B9:J9')
    ws4.append([])
    
    ws4.append(["【界面截图说明】"])
    ws4.merge_cells('A11:J11')
    ws4['A11'].font = Font(bold=True, color='C00000', size=11)
    ws4.append([])
    ws4.append(["图片引用", "ui-screenshot.png (位于同目录)"])
    ws4.merge_cells('B13:J13')
    ws4.append(["截图内容", "编钟模拟器实际运行界面，展示完整的三排编钟布局和UI元素"])
    ws4.merge_cells('B14:J14')
    ws4.append([])
    
    ws4.append(["UI布局详解"])
    ws4.merge_cells('A16:J16')
    ws4['A16'].font = subtitle_font
    ws4.append([])
    ws4.append(["区域", "内容说明", "设计细节"])
    ws4.merge_cells('C18:J18')
    ws4['A18'].font = Font(bold=True)
    ws4['B18'].font = Font(bold=True)
    ws4['C18'].font = Font(bold=True)
    ws4.append(["顶部工具栏", "应用名称+录制/我的演奏/设计案按钮", "悬浮透明背景，避免遮挡主体，右侧按钮采用胶囊设计"])
    ws4.merge_cells('C19:J19')
    ws4.append(["状态栏", "系统状态+录制计时", "录制时红色指示灯闪烁，提供视觉反馈"])
    ws4.merge_cells('C20:J20')
    ws4.append(["编钟主体区", "三排15口编钟", "Flexbox垂直布局，钟架使用木质渐变(#8B4513→#654321)增加层次"])
    ws4.merge_cells('C21:J21')
    ws4.append(["底部控制栏", "试听/清空/回放按钮", "圆角胶囊按钮，符合移动端触控习惯，间距16px"])
    ws4.merge_cells('C22:J22')
    ws4.append([])
    
    ws4.append(["【界面结构示意图】"])
    ws4.merge_cells('A24:J24')
    ws4['A24'].font = Font(bold=True, color='C00000', size=11)
    ws4.append([])
    diagram_text = """┌─────────────────────────────────────┐
│  🔧 工具栏                          │  ← 标题 + 录制按钮
│  [🔔 编钟模拟器]        [🔴 录制]   │
├─────────────────────────────────────┤
│  📊 状态栏                          │  ← 状态提示 + 录制计时
│  ● 准备就绪          ⏱️ 00:00       │
├─────────────────────────────────────┤
│                                     │
│         🏛️ 编钟区域                  │  ← 核心交互区
│      ┌───────────────┐              │
│      │ ○ ○ ○ ○ ○   │  上层(高音)   │
│      │ ○  ○  ○  ○  ○ │  中层(中音)   │
│      │ ○   ○   ○   ○   ○ │  下层(低音)   │
│      └───────────────┘              │
│                                     │
├─────────────────────────────────────┤
│  🎮 控制面板                        │  ← 快捷操作
│  [▶️ 试听] [🗑️ 清空] [▶️ 回放]      │
└─────────────────────────────────────┘"""
    ws4.append([diagram_text])
    ws4.merge_cells('A25:J25')
    ws4['A25'].font = code_font
    ws4['A25'].alignment = Alignment(wrap_text=True, vertical='top')
    ws4.row_dimensions[25].height = 200
    ws4.append([])
    
    ws4.append(["UE设计亮点"])
    ws4.merge_cells('A27:J27')
    ws4['A27'].font = subtitle_font
    ws4.append([])
    ws4.append(["亮点", "说明"])
    ws4.merge_cells('B29:J29')
    ws4['A29'].font = Font(bold=True)
    ws4['B29'].font = Font(bold=True)
    ws4.append(["沉浸优先", "深色渐变背景(#1a1a2e→#16213e→#0f3460)将视觉焦点完全引导至金色编钟"])
    ws4.merge_cells('B30:J30')
    ws4.append(["直觉交互", "编钟点击区域覆盖整个钟体，左右半区分别对应侧敲/正敲，无需文字说明"])
    ws4.merge_cells('B31:J31')
    ws4.append(["即时反馈", "每次点击触发钟体摆动动画(±5°/0.3s) + Toast提示 + 音效，三重感官确认"])
    ws4.merge_cells('B32:J32')
    ws4.append(["横屏强制", "竖屏时显示旋转提示，确保编钟的宽展布局得到最佳展示"])
    ws4.merge_cells('B33:J33')
    ws4.append([])
    
    ws4.append(["色彩系统"])
    ws4.merge_cells('A35:J35')
    ws4['A35'].font = subtitle_font
    ws4.append([])
    ws4.append(["元素", "颜色值", "用途"])
    ws4.merge_cells('C37:J37')
    ws4['A37'].font = Font(bold=True)
    ws4['B37'].font = Font(bold=True)
    ws4['C37'].font = Font(bold=True)
    ws4.append(["背景", "#1a1a2e → #16213e → #0f3460", "深蓝渐变，营造庄重氛围"])
    ws4.merge_cells('C38:J38')
    ws4.append(["编钟", "#FFD700 → #DAA520 → #B8860B", "金色渐变，青铜质感"])
    ws4.merge_cells('C39:J39')
    ws4.append(["钟架", "#8B4513 → #654321", "木质纹理"])
    ws4.merge_cells('C40:J40')
    ws4.append(["强调色", "#f5af19", "标题、按钮、高亮"])
    ws4.merge_cells('C41:J41')
    ws4.append(["录制中", "#ff0000", "闪烁动画表示录制状态"])
    ws4.merge_cells('C42:J42')
    ws4.append([])
    
    ws4.append(["视觉反馈系统"])
    ws4.merge_cells('A44:J44')
    ws4['A44'].font = subtitle_font
    ws4.append([])
    ws4.append(["操作", "视觉反馈", "动画效果"])
    ws4.merge_cells('C46:J46')
    ws4['A46'].font = Font(bold=True)
    ws4['B46'].font = Font(bold=True)
    ws4['C46'].font = Font(bold=True)
    ws4.append(["点击编钟", "钟体摆动", "0.3s摇摆动画(±5°)，ease-out缓动"])
    ws4.merge_cells('C47:J47')
    ws4.append(["正敲/侧敲", "击槌图标显示", "锤击动画+位置提示(左侧显示「侧」右侧显示「正」)"])
    ws4.merge_cells('C48:J48')
    ws4.append(["开始录制", "录制按钮变红+脉冲动画", "1s闪烁周期，红色(#ff0000)呼吸效果"])
    ws4.merge_cells('C49:J49')
    ws4.append(["Toast提示", "底部弹出提示", "滑入滑出，2s显示，半透明黑色背景"])
    ws4.merge_cells('C50:J50')
    
    # 设置列宽
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        for col in ['C','D','E','F','G','H','I','J']:
            ws.column_dimensions[col].width = 15
    
    # 保存
    output_path = '/root/.openclaw/workspace/编钟模拟器设计案_完整版.xlsx'
    wb.save(output_path)
    print(f"✅ Excel文件已生成: {output_path}")
    return output_path

if __name__ == '__main__':
    create_complete_design_doc()
