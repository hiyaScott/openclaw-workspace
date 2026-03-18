#!/usr/bin/env python3
"""
编钟模拟器设计案 - Excel 导出
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_bianzhong_design_doc():
    wb = openpyxl.Workbook()
    
    # 删除默认sheet，按顺序创建各个部分
    wb.remove(wb.active)
    
    # ===== Sheet 1: 项目概述 =====
    ws1 = wb.create_sheet("1.项目概述")
    ws1.append(["项目概述", "", ""])
    ws1.append([])
    ws1.append(["项目名称", "编钟模拟器", ""])
    ws1.append(["项目类型", "文化体验型音乐应用 / 音游原型", ""])
    ws1.append(["灵感来源", "曾侯乙编钟（战国早期，约公元前433年）", ""])
    ws1.append(["核心特色", "一钟双音系统 - 每口钟可发出正敲/侧敲两个不同音高", ""])
    ws1.append([])
    ws1.append(["设计目标", "", ""])
    ws1.append(["1", "在移动端浏览器实现接近原生 App 的触控体验", ""])
    ws1.append(["2", "准确还原编钟「一钟双音」的声学特性", ""])
    ws1.append(["3", "提供完整的录制-回放-分享创作闭环", ""])
    ws1.append(["4", "保持界面简洁优雅，突出编钟本身的美学", ""])
    
    # ===== Sheet 2: 核心机制 =====
    ws2 = wb.create_sheet("2.核心机制")
    ws2.append(["核心机制", "", "", ""])
    ws2.append([])
    ws2.append(["一钟双音系统", "", "", ""])
    ws2.append(["敲击方式", "触发位置", "音高关系", "音色特点"])
    ws2.append(["正敲（正鼓音）", "钟体正中", "基音", "浑厚、低沉"])
    ws2.append(["侧敲（侧鼓音）", "钟体两侧", "基音 + 大三度（+4半音）", "清亮、通透"])
    ws2.append([])
    ws2.append(["音区架构", "", "", ""])
    ws2.append(["排层", "钟大小", "音区范围", "数量", "视觉设计"])
    ws2.append(["上层", "小编钟", "C5 - A5（高音）", "5口", "尺寸最小，位置最高"])
    ws2.append(["中层", "中编钟", "C4 - A4（中音）", "5口", "中等尺寸，中间位置"])
    ws2.append(["下层", "大编钟", "C3 - A3（低音）", "5口", "尺寸最大，位置最低"])
    ws2.append([])
    ws2.append(["总计", "15口编钟 × 2音/钟 = 30个可用音高", "", "", ""])
    
    # ===== Sheet 3: 音频配置表 =====
    ws3 = wb.create_sheet("3.音频配置")
    ws3.append(["音频配置表", "", "", "", ""])
    ws3.append([])
    ws3.append(["上层（高音）", "", "", "", ""])
    ws3.append(["音名", "正敲频率(Hz)", "侧敲频率(Hz)", "音分差", "波形类型"])
    ws3.append(["C5", "523.25", "659.25", "400", "Sine + Triangle"])
    ws3.append(["D5", "587.33", "698.46", "300", "Sine + Triangle"])
    ws3.append(["E5", "659.25", "783.99", "300", "Sine + Triangle"])
    ws3.append(["G5", "783.99", "987.77", "400", "Sine + Triangle"])
    ws3.append(["A5", "880.00", "1046.50", "300", "Sine + Triangle"])
    ws3.append([])
    ws3.append(["中层（中音）", "", "", "", ""])
    ws3.append(["音名", "正敲频率(Hz)", "侧敲频率(Hz)", "音分差", "波形类型"])
    ws3.append(["C4", "261.63", "329.63", "400", "Sine + Triangle"])
    ws3.append(["D4", "293.66", "349.23", "300", "Sine + Triangle"])
    ws3.append(["E4", "329.63", "392.00", "300", "Sine + Triangle"])
    ws3.append(["G4", "392.00", "493.88", "400", "Sine + Triangle"])
    ws3.append(["A4", "440.00", "523.25", "300", "Sine + Triangle"])
    ws3.append([])
    ws3.append(["下层（低音）", "", "", "", ""])
    ws3.append(["音名", "正敲频率(Hz)", "侧敲频率(Hz)", "音分差", "波形类型"])
    ws3.append(["C3", "130.81", "164.81", "400", "Sine + Triangle"])
    ws3.append(["D3", "146.83", "174.61", "300", "Sine + Triangle"])
    ws3.append(["E3", "164.81", "196.00", "300", "Sine + Triangle"])
    ws3.append(["G3", "196.00", "246.94", "400", "Sine + Triangle"])
    ws3.append(["A3", "220.00", "261.63", "300", "Sine + Triangle"])
    
    # ===== Sheet 4: 音色合成 =====
    ws4 = wb.create_sheet("4.音色合成")
    ws4.append(["音色合成原理", "", "", "", ""])
    ws4.append([])
    ws4.append(["层次", "波形类型", "频率", "音量", "衰减时间", "作用"])
    ws4.append(["基频层", "正弦波", "f（基频）", "0.6", "3秒", "主体音色"])
    ws4.append(["二次谐波", "正弦波", "2f（八度）", "0.3", "2.4秒", "丰富度"])
    ws4.append(["三次谐波", "三角波", "3f", "0.15", "1.5秒", "金属质感"])
    ws4.append([])
    ws4.append(["包络设计", "", "", "", ""])
    ws4.append(["参数", "数值", "说明", "", ""])
    ws4.append(["Attack", "20ms", "快速起音", "", ""])
    ws4.append(["Decay", "3000ms", "长衰减，模拟铜钟持续振动", "", ""])
    ws4.append([])
    ws4.append(["音频路由", "", "", "", ""])
    ws4.append(["步骤", "节点", "配置", "", ""])
    ws4.append(["1", "振荡器组", "3个振荡器叠加", "", ""])
    ws4.append(["2", "增益节点", "Volume: 0.6", "", ""])
    ws4.append(["3", "混响节点", "ConvolverNode模拟大型音乐厅", "", ""])
    ws4.append(["4", "主增益", "Volume: 0.8", "", ""])
    ws4.append(["5", "音频输出", "destination", "", ""])
    
    # ===== Sheet 5: UI/UE设计 =====
    ws5 = wb.create_sheet("5.UI-UE设计")
    ws5.append(["UI/UE设计", "", "", ""])
    ws5.append([])
    ws5.append(["设计原则", "", "", ""])
    ws5.append(["原则", "说明", "", ""])
    ws5.append(["沉浸优先", "深色背景减少视觉干扰，突出金色编钟", "", ""])
    ws5.append(["横屏体验", "强制横屏模式，模拟真实编钟的宽展布局", "", ""])
    ws5.append(["直观操作", "无需教程，点击/触控即可发声", "", ""])
    ws5.append(["即时反馈", "每次敲击都有视觉和听觉双重反馈", "", ""])
    ws5.append([])
    ws5.append(["界面布局", "", "", ""])
    ws5.append(["区域", "内容", "设计说明", ""])
    ws5.append(["顶部工具栏", "应用名称+录制/我的演奏/设计案按钮", "悬浮透明背景，避免遮挡主体", ""])
    ws5.append(["状态栏", "系统状态+录制计时", "录制时红色指示灯闪烁", ""])
    ws5.append(["编钟主体区", "三排15口编钟", "Flexbox垂直布局，钟架木质渐变", ""])
    ws5.append(["底部控制栏", "试听/清空/回放按钮", "圆角胶囊按钮，适合移动端触控", ""])
    ws5.append([])
    ws5.append(["色彩系统", "", "", ""])
    ws5.append(["元素", "颜色值", "用途", ""])
    ws5.append(["背景", "#1a1a2e → #16213e → #0f3460", "深蓝渐变，庄重氛围", ""])
    ws5.append(["编钟", "#FFD700 → #DAA520 → #B8860B", "金色渐变，青铜质感", ""])
    ws5.append(["钟架", "#8B4513 → #654321", "木质纹理", ""])
    ws5.append(["强调色", "#f5af19", "标题、按钮、高亮", ""])
    ws5.append(["录制中", "#ff0000", "闪烁动画表示录制状态", ""])
    
    # ===== Sheet 6: 交互逻辑 =====
    ws6 = wb.create_sheet("6.交互逻辑")
    ws6.append(["交互逻辑", "", "", ""])
    ws6.append([])
    ws6.append(["敲击检测机制", "", "", ""])
    ws6.append(["点击位置", "触发方式", "音高", ""])
    ws6.append(["钟体左半区 (x < width/2)", "侧敲", "侧鼓音（高音）", ""])
    ws6.append(["钟体右半区 (x ≥ width/2)", "正敲", "正鼓音（低音）", ""])
    ws6.append([])
    ws6.append(["视觉反馈系统", "", "", ""])
    ws6.append(["操作", "视觉反馈", "动画效果", ""])
    ws6.append(["点击编钟", "钟体摆动", "0.3s摇摆动画(±5°)", ""])
    ws6.append(["正敲/侧敲", "击槌图标显示", "锤击动画+位置提示", ""])
    ws6.append(["开始录制", "录制按钮变红+脉冲动画", "1s闪烁周期", ""])
    ws6.append(["Toast提示", "底部弹出提示", "滑入滑出，2s显示", ""])
    ws6.append([])
    ws6.append(["屏幕方向处理", "", "", ""])
    ws6.append(["状态", "处理方式", "", ""])
    ws6.append(["横屏", "正常显示编钟界面", "", ""])
    ws6.append(["竖屏", "显示旋转提示：「请横屏使用编钟模拟器」", "隐藏主界面，提供旋转引导动画", ""])
    
    # ===== Sheet 7: 录制回放系统 =====
    ws7 = wb.create_sheet("7.录制回放")
    ws7.append(["录制与回放系统", "", "", "", ""])
    ws7.append([])
    ws7.append(["数据结构", "", "", "", ""])
    ws7.append(["字段", "类型", "说明", "示例", ""])
    ws7.append(["timestamp", "Number(ms)", "相对于录制开始的毫秒时间戳", "0, 500, 1200", ""])
    ws7.append(["bellId", "String", "编钟标识，如 upper-1", "upper-1, middle-3", ""])
    ws7.append(["hitType", "String", "敲击类型：正敲或侧敲", "正敲, 侧敲", ""])
    ws7.append(["frequency", "Number", "实际播放频率(Hz)", "261.63, 329.63", ""])
    ws7.append([])
    ws7.append(["录制流程", "", "", "", ""])
    ws7.append(["步骤", "操作", "说明", "", ""])
    ws7.append(["1", "点击录制按钮", "开始记录，状态栏显示红色录制指示", "", ""])
    ws7.append(["2", "演奏编钟", "所有敲击事件被记录到事件数组", "", ""])
    ws7.append(["3", "停止录制", "计算总时长，生成元数据", "", ""])
    ws7.append(["4", "保存", "数据存入LocalStorage，可导出JSON", "", ""])
    ws7.append([])
    ws7.append(["回放机制", "", "", "", ""])
    ws7.append(["技术方案", "setTimeout链式触发", "", "", ""])
    ws7.append(["精度", "毫秒级", "", "", ""])
    ws7.append(["控制功能", "播放/暂停/停止/进度拖动", "", "", ""])
    ws7.append([])
    ws7.append(["分享功能", "", "", "", ""])
    ws7.append(["方式", "说明", "", "", ""])
    ws7.append(["链接分享", "生成包含data参数URL", "", "", ""])
    ws7.append(["图片分享", "Canvas生成演奏截图", "", "", ""])
    ws7.append(["JSON导出", "原始数据文件下载", "", "", ""])
    
    # ===== Sheet 8: 技术实现 =====
    ws8 = wb.create_sheet("8.技术实现")
    ws8.append(["技术实现", "", "", ""])
    ws8.append([])
    ws8.append(["技术栈", "", "", ""])
    ws8.append(["类别", "技术", "用途", ""])
    ws8.append(["音频合成", "Web Audio API", "实时生成编钟音色", ""])
    ws8.append(["数据存储", "localStorage", "录制数据持久化", ""])
    ws8.append(["图形渲染", "Canvas API", "分享图片生成", ""])
    ws8.append(["界面框架", "原生HTML/CSS/JS", "无框架依赖，轻量快速", ""])
    ws8.append([])
    ws8.append(["浏览器兼容性", "", "", ""])
    ws8.append(["浏览器", "支持情况", "注意事项", ""])
    ws8.append(["Chrome/Edge", "✅ 完全支持", "推荐浏览器", ""])
    ws8.append(["Safari", "✅ 支持", "iOS需用户手势触发音频", ""])
    ws8.append(["Firefox", "✅ 支持", "", ""])
    ws8.append([])
    ws8.append(["音频自动播放策略", "", "", ""])
    ws8.append(["平台", "策略", "解决方案", ""])
    ws8.append(["桌面端", "无需用户交互", "直接播放", ""])
    ws8.append(["iOS Safari", "需要用户手势触发", "首次点击解除静音", ""])
    ws8.append(["Android Chrome", "需要用户交互", "首次点击解除静音", ""])
    
    # ===== Sheet 9: 开发路线图 =====
    ws9 = wb.create_sheet("9.开发路线")
    ws9.append(["开发路线图", "", "", "", ""])
    ws9.append([])
    ws9.append(["已实现功能", "", "", "", ""])
    ws9.append(["模块", "功能", "状态", "", ""])
    ws9.append(["音频系统", "一钟双音合成", "✅ 完成", "", ""])
    ws9.append(["音频系统", "三层振荡器音色", "✅ 完成", "", ""])
    ws9.append(["音频系统", "混响效果", "✅ 完成", "", ""])
    ws9.append(["交互", "点击/触摸敲击", "✅ 完成", "", ""])
    ws9.append(["交互", "正敲/侧敲判定", "✅ 完成", "", ""])
    ws9.append(["交互", "视觉反馈动画", "✅ 完成", "", ""])
    ws9.append(["录制", "事件记录", "✅ 完成", "", ""])
    ws9.append(["录制", "LocalStorage存储", "✅ 完成", "", ""])
    ws9.append(["回放", "setTimeout触发", "✅ 完成", "", ""])
    ws9.append(["回放", "播放控制", "✅ 完成", "", ""])
    ws9.append(["分享", "JSON导出", "✅ 完成", "", ""])
    ws9.append(["分享", "链接分享", "✅ 完成", "", ""])
    ws9.append([])
    ws9.append(["未来扩展", "", "", "", ""])
    ws9.append(["功能", "说明", "优先级", "", ""])
    ws9.append(["节拍器", "可视化节拍引导", "中", "", ""])
    ws9.append(["多编钟切换", "曾侯乙/长台关等不同形制", "低", "", ""])
    ws9.append(["教程模式", "引导式教学", "中", "", ""])
    ws9.append(["社交功能", "作品点赞评论", "低", "", ""])
    ws9.append(["音游模式", "曲谱挑战+评分", "高", "", ""])
    ws9.append(["AI作曲", "智能伴奏生成", "低", "", ""])
    
    # ===== Sheet 10: 音游化扩展 =====
    ws10 = wb.create_sheet("10.音游化扩展")
    ws10.append(["音游化扩展设计", "", "", "", ""])
    ws10.append([])
    ws10.append(["特色玩法方向", "", "", "", ""])
    ws10.append(["方向", "玩法说明", "特色", ""])
    ws10.append(["和声解密", "屏幕上显示和弦名称，玩家需同时按下多个钟的正/侧面拼出和弦", "别的音游没有「一钟双音」", ""])
    ws10.append(["古乐还原", "播放编钟实际演奏录音，玩家跟随演奏", "真实历史乐器跟弹体验", ""])
    ws10.append(["对位竞技", "双人分屏，各控制一半音区合奏", "天然适合分工协作", ""])
    ws10.append([])
    ws10.append(["谱面可视化方案", "", "", "", ""])
    ws10.append(["方案", "描述", "", ""])
    ws10.append(["分层轨道", "上层=正敲音符(实心圆)，下层=侧敲音符(空心圆)", "", ""])
    ws10.append(["钟体可视化", "谱面即编钟俯视图，音符落入对应钟位置", "", ""])
    ws10.append([])
    ws10.append(["难度设计", "", "", "", ""])
    ws10.append(["难度", "机制", "示例", ""])
    ws10.append(["Easy", "单音旋律，只用到正敲", "《小星星》", ""])
    ws10.append(["Normal", "简单和弦，偶尔双音", "三和弦分解", ""])
    ws10.append(["Hard", "密集双音，跨区快速切换", "巴赫风格对位", ""])
    ws10.append(["Master", "复节奏，全音区同时触发", "完整交响编配", ""])
    ws10.append([])
    ws10.append(["UGC生态设计", "", "", "", ""])
    ws10.append(["环节", "说明", "", ""])
    ws10.append(["创作者演奏", "高端玩家自由演奏/即兴创作", "", ""])
    ws10.append(["生成谱面", "系统自动量化时间戳、标定难度", "", ""])
    ws10.append(["分享分发", "谱面+预览音频+热力图缩略", "", ""])
    ws10.append(["玩家挑战", "普通玩家游玩获得评分", "", ""])
    ws10.append(["数据反馈", "通关率、retry分布等反馈给创作者", "", ""])
    
    # 设置所有sheet的样式
    for ws in wb.worksheets:
        # 标题行样式
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.row_dimensions[1].height = 30
        
        # 表头样式（如果第3行有内容且是表头）
        if ws.max_row >= 3:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=3, column=col)
                if cell.value and any(keyword in str(cell.value) for keyword in ['字段', '层级', '步骤', '模块', '音名', '方向']):
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        
        # 第一列加宽（标题列）
        ws.column_dimensions['A'].width = 18
        
        # 自动调整行高
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
    
    # 保存文件
    output_path = '/root/.openclaw/workspace/编钟模拟器设计案.xlsx'
    wb.save(output_path)
    print(f"✅ Excel文件已生成: {output_path}")
    
    # 生成文件信息
    import os
    file_size = os.path.getsize(output_path)
    print(f"📊 文件大小: {file_size / 1024:.1f} KB")
    print(f"📑 包含工作表: {len(wb.worksheets)} 个")
    for i, ws in enumerate(wb.worksheets, 1):
        print(f"   {i}. {ws.title}")
    
    return output_path

if __name__ == '__main__':
    create_bianzhong_design_doc()
