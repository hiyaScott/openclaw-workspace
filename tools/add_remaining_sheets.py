#!/usr/bin/env python3
"""
编钟模拟器设计案 - 完整版 Excel 导出（续）
添加剩余工作表
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def add_remaining_sheets():
    # 加载已创建的文件
    wb = openpyxl.load_workbook('/root/.openclaw/workspace/编钟模拟器设计案_完整版.xlsx')
    
    # 定义样式
    title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    subtitle_font = Font(name='微软雅黑', size=11, bold=True, color='1F4E79')
    note_font = Font(name='微软雅黑', size=9, color='666666', italic=True)
    code_font = Font(name='Courier New', size=9)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    def style_header(ws, title):
        ws.append([title])
        cell = ws.cell(row=1, column=1)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.row_dimensions[1].height = 35
        ws.append([])
    
    # ===== Sheet 5: 交互逻辑 =====
    ws5 = wb.create_sheet("5.交互逻辑")
    style_header(ws5, "👆 交互逻辑")
    
    ws5.append(["【点击区域划分示意图】"])
    ws5.merge_cells('A3:J3')
    ws5['A3'].font = Font(bold=True, color='C00000', size=11)
    ws5.append([])
    
    hit_diagram = """┌─────────────────────────┐
│       编钟钟体           │
│   ◄─────┼─────►         │
│   侧敲  │  正敲          │
│ (左侧)  │ (右侧)         │
│   E4    │   C4           │
│  高音   │  低音          │
└─────────────────────────┘

判定逻辑：
if (点击位置.x < 钟体宽度 / 2) → 侧敲 (侧鼓音)
else → 正敲 (正鼓音)"""
    ws5.append([hit_diagram])
    ws5.merge_cells('A4:J4')
    ws5['A4'].font = code_font
    ws5['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws5.row_dimensions[4].height = 200
    ws5.append([])
    
    ws5.append(["敲击检测机制"])
    ws5.merge_cells('A6:J6')
    ws5['A6'].font = subtitle_font
    ws5.append([])
    ws5.append(["检测方式", "点击位置判断"])
    ws5.merge_cells('B8:J8')
    ws5.append(["悬停提示", "鼠标/手指悬停时显示「正」和「侧」字样，帮助用户理解交互方式"])
    ws5.merge_cells('B9:J9')
    ws5.append([])
    
    ws5.append(["手势支持"])
    ws5.merge_cells('A11:J11')
    ws5['A11'].font = subtitle_font
    ws5.append([])
    ws5.append(["手势", "对应操作", "触发方式"])
    ws5.merge_cells('D13:J13')
    ws5['A13'].font = Font(bold=True)
    ws5['B13'].font = Font(bold=True)
    ws5['C13'].font = Font(bold=True)
    ws5.append(["点击/触摸", "敲击编钟", "click / touchstart事件"])
    ws5.merge_cells('D14:J14')
    ws5.append(["横滑", "快速连续敲击", "多点触控支持，滑动经过多钟触发连击"])
    ws5.merge_cells('D15:J15')
    ws5.append([])
    ws5.append(["防误触处理"])
    ws5.merge_cells('A17:J17')
    ws5['A17'].font = Font(bold=True)
    ws5.append(["-webkit-tap-highlight-color: transparent", "消除移动端点击高亮"])
    ws5.merge_cells('B18:J18')
    ws5.append([])
    
    ws5.append(["屏幕方向处理"])
    ws5.merge_cells('A20:J20')
    ws5['A20'].font = subtitle_font
    ws5.append([])
    ws5.append(["状态", "处理方式"])
    ws5.merge_cells('C22:J22')
    ws5['A22'].font = Font(bold=True)
    ws5['B22'].font = Font(bold=True)
    ws5.append(["横屏", "正常显示编钟界面"])
    ws5.merge_cells('C23:J23')
    ws5.append(["竖屏", "显示旋转提示：「请横屏使用编钟模拟器」，隐藏主界面，提供旋转引导动画"])
    ws5.merge_cells('C24:J24')
    ws5.append([])
    ws5.append(["实现方式", "@media screen and (orientation: portrait) CSS媒体查询"])
    ws5.merge_cells('B26:J26')
    
    # ===== Sheet 6: 录制与回放 =====
    ws6 = wb.create_sheet("6.录制与回放")
    style_header(ws6, "⏺️ 录制与回放系统")
    
    ws6.append(["【数据结构设计 - Recording对象】"])
    ws6.merge_cells('A3:J3')
    ws6['A3'].font = Font(bold=True, color='C00000', size=11)
    ws6.append([])
    
    data_struct = """{
    id: 1709999999999,           // 时间戳作为唯一ID
    name: "演奏 1",              // 用户可编辑名称
    date: "2026/3/11 10:30:00",  // 录制日期时间
    duration: "45.2",            // 演奏时长（秒）
    notes: [                     // 音符数组
        { note: "C4", freq: 261.63, isSide: false, time: 0 },
        { note: "E4", freq: 329.63, isSide: true, time: 850 },
        // ...
    ]
}"""
    ws6.append([data_struct])
    ws6.merge_cells('A5:J5')
    ws6['A5'].font = code_font
    ws6['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws6.row_dimensions[5].height = 150
    ws6.append([])
    
    ws6.append(["录制流程"])
    ws6.merge_cells('A7:J7')
    ws6['A7'].font = subtitle_font
    ws6.append([])
    ws6.append(["步骤", "操作", "说明"])
    ws6.merge_cells('D9:J9')
    ws6['A9'].font = Font(bold=True)
    ws6['B9'].font = Font(bold=True)
    ws6['C9'].font = Font(bold=True)
    ws6.append(["1", "点击录制按钮", "开始记录，状态栏显示红色录制指示"])
    ws6.merge_cells('D10:J10')
    ws6.append(["2", "初始化音频上下文", "确保AudioContext处于running状态"])
    ws6.merge_cells('D11:J11')
    ws6.append(["3", "开始计时", "记录起始时间戳"])
    ws6.merge_cells('D12:J12')
    ws6.append(["4", "演奏编钟", "所有敲击事件被记录到事件数组（时间+音高+方式）"])
    ws6.merge_cells('D13:J13')
    ws6.append(["5", "停止录制", "计算总时长，生成元数据"])
    ws6.merge_cells('D14:J14')
    ws6.append(["6", "保存", "数据存入LocalStorage，可导出JSON"])
    ws6.merge_cells('D15:J15')
    ws6.append([])
    
    ws6.append(["回放机制"])
    ws6.merge_cells('A17:J17')
    ws6['A17'].font = subtitle_font
    ws6.append([])
    ws6.append(["技术方案", "setTimeout按照录制时的时间戳精确触发每个音符"])
    ws6.merge_cells('B19:J19')
    ws6.append([])
    code_replay = """recording.notes.forEach((noteData) => {
    setTimeout(() => {
        playBellSound(noteData.freq);
    }, noteData.time);
});"""
    ws6.append(["代码示例:"])
    ws6.merge_cells('A20:J20')
    ws6.append([code_replay])
    ws6.merge_cells('A21:J21')
    ws6['A21'].font = code_font
    ws6['A21'].alignment = Alignment(wrap_text=True, vertical='top')
    ws6.row_dimensions[21].height = 60
    ws6.append([])
    ws6.append(["特点", "非音频文件录制(数据体积极小)、精确还原演奏时机、支持变速回放(可扩展)"])
    ws6.merge_cells('B23:J23')
    ws6.append([])
    
    ws6.append(["分享功能"])
    ws6.merge_cells('A25:J25')
    ws6['A25'].font = subtitle_font
    ws6.append([])
    ws6.append(["分享方式", "实现方式", "数据格式"])
    ws6.merge_cells('D27:J27')
    ws6['A27'].font = Font(bold=True)
    ws6['B27'].font = Font(bold=True)
    ws6['C27'].font = Font(bold=True)
    ws6.append(["链接分享", "Base64编码 + URL参数", "?share=eyJuYW1lIjo..."])
    ws6.merge_cells('D28:J28')
    ws6.append(["图片卡片", "Canvas绘制 + 下载", "PNG图片"])
    ws6.merge_cells('D29:J29')
    ws6.append(["数据导出", "Blob + 文件下载", "JSON文件"])
    ws6.merge_cells('D30:J30')
    ws6.append([])
    
    ws6.append(["存储方案"])
    ws6.merge_cells('A32:J32')
    ws6['A32'].font = subtitle_font
    ws6.append([])
    ws6.append(["方案", "LocalStorage持久化"])
    ws6.merge_cells('B34:J34')
    ws6.append(["键名", "bianzhong_recordings"])
    ws6.merge_cells('B35:J35')
    ws6.append(["格式", "JSON字符串"])
    ws6.merge_cells('B36:J36')
    ws6.append(["容量", "浏览器限制约5-10MB"])
    ws6.merge_cells('B37:J37')
    ws6.append(["数据量估算", "一首30秒演奏约500个音符 ≈ 10KB"])
    ws6.merge_cells('B38:J38')
    ws6.append(["理论容量", "约500-1000首演奏"])
    ws6.merge_cells('B39:J39')
    
    # ===== Sheet 7: 技术实现 =====
    ws7 = wb.create_sheet("7.技术实现")
    style_header(ws7, "💻 技术实现")
    
    ws7.append(["技术栈"])
    ws7.merge_cells('A3:J3')
    ws7['A3'].font = subtitle_font
    ws7.append([])
    ws7.append(["类别", "技术", "版本/说明"])
    ws7.merge_cells('D5:J5')
    ws7['A5'].font = Font(bold=True)
    ws7['B5'].font = Font(bold=True)
    ws7['C5'].font = Font(bold=True)
    ws7.append(["核心API", "Web Audio API", "原生浏览器支持"])
    ws7.merge_cells('D6:J6')
    ws7.append(["存储", "localStorage", "录制数据持久化"])
    ws7.merge_cells('D7:J7')
    ws7.append(["绘图", "Canvas API", "分享卡片生成"])
    ws7.merge_cells('D8:J8')
    ws7.append(["样式", "CSS3", "Flexbox + 渐变 + 动画"])
    ws7.merge_cells('D9:J9')
    ws7.append(["框架", "无", "纯原生实现"])
    ws7.merge_cells('D10:J10')
    ws7.append([])
    
    ws7.append(["浏览器兼容性"])
    ws7.merge_cells('A12:J12')
    ws7['A12'].font = subtitle_font
    ws7.append([])
    ws7.append(["浏览器", "支持情况", "备注"])
    ws7.merge_cells('D14:J14')
    ws7['A14'].font = Font(bold=True)
    ws7['B14'].font = Font(bold=True)
    ws7['C14'].font = Font(bold=True)
    ws7.append(["Chrome/Edge", "✅ 完全支持", "推荐浏览器"])
    ws7.merge_cells('D15:J15')
    ws7.append(["Firefox", "✅ 完全支持", ""])
    ws7.merge_cells('D16:J16')
    ws7.append(["Safari", "✅ 完全支持", "iOS需用户交互激活音频"])
    ws7.merge_cells('D17:J17')
    ws7.append(["微信内置浏览器", "⚠️ 部分支持", "可能有自动播放限制"])
    ws7.merge_cells('D18:J18')
    ws7.append([])
    
    ws7.append(["音频自动播放策略"])
    ws7.merge_cells('A20:J20')
    ws7['A20'].font = subtitle_font
    ws7.append([])
    ws7.append(["现代浏览器对自动播放有严格限制，采用以下策略处理："])
    ws7.merge_cells('A22:J22')
    ws7.append([])
    code_audio = """function initAudio() {
    if (!audioContext) {
        audioContext = new (window.AudioContext || window.webkitAudioContext)();
        masterGain = audioContext.createGain();
        masterGain.gain.value = 0.8;
        masterGain.connect(audioContext.destination);
    }
    // 处理暂停状态（iOS Safari需要）
    if (audioContext.state === 'suspended') {
        audioContext.resume();
    }
}

// 在用户首次点击/触摸时初始化
bell.addEventListener('click', () => {
    initAudio();
    playBellSound(freq);
});"""
    ws7.append([code_audio])
    ws7.merge_cells('A24:J24')
    ws7['A24'].font = code_font
    ws7['A24'].alignment = Alignment(wrap_text=True, vertical='top')
    ws7.row_dimensions[24].height = 180
    
    # ===== Sheet 8: 开发路线 =====
    ws8 = wb.create_sheet("8.开发路线")
    style_header(ws8, "🗺️ 开发路线图")
    
    ws8.append(["已实现功能"])
    ws8.merge_cells('A3:J3')
    ws8['A3'].font = subtitle_font
    ws8['A3'].font = Font(bold=True, color='00B050')
    ws8.append([])
    ws8.append(["标签", "功能", "状态"])
    ws8.merge_cells('D5:J5')
    ws8['A5'].font = Font(bold=True)
    ws8['B5'].font = Font(bold=True)
    ws8['C5'].font = Font(bold=True)
    
    features = [
        ("核心", "一钟双音系统（正敲/侧敲）"),
        ("核心", "三排15口编钟音阶布局"),
        ("核心", "Web Audio API音色合成"),
        ("UI", "响应式横屏布局"),
        ("UI", "编钟摆动动画"),
        ("技术", "录制与回放系统"),
        ("技术", "LocalStorage持久化存储"),
        ("技术", "分享链接生成"),
        ("技术", "分享卡片生成（Canvas）"),
        ("技术", "数据导出为JSON")
    ]
    for tag, feat in features:
        ws8.append([tag, feat, "✅ 已完成"])
        ws8.merge_cells(f'D{ws8.max_row}:J{ws8.max_row}')
    
    ws8.append([])
    ws8.append(["未来可扩展功能"])
    ws8.merge_cells(f'A{ws8.max_row}:J{ws8.max_row}')
    ws8.cell(row=ws8.max_row, column=1).font = subtitle_font
    ws8.append([])
    ws8.append(["功能", "优先级", "说明"])
    ws8.merge_cells('D{0}:J{0}'.format(ws8.max_row))
    ws8.cell(row=ws8.max_row, column=1).font = Font(bold=True)
    ws8.cell(row=ws8.max_row, column=2).font = Font(bold=True)
    ws8.cell(row=ws8.max_row, column=3).font = Font(bold=True)
    
    future = [
        ("更多编钟数量", "P1", "增加至21口或25口，扩展音域"),
        ("节拍器", "P1", "辅助演奏者保持节奏"),
        ("音游模式", "P1", "曲谱挑战+评分系统"),
        ("录音重命名", "P2", "支持自定义演奏名称"),
        ("变速回放", "P2", "0.5x-2x速度调节"),
        ("音轨叠加", "P2", "多轨录制，类似Overdub"),
        ("教程模式", "P3", "引导新手学习演奏简单曲目"),
        ("云端存储", "P3", "用户账号系统，跨设备同步")
    ]
    for feat, priority, desc in future:
        ws8.append([feat, priority, desc])
        ws8.merge_cells(f'D{ws8.max_row}:J{ws8.max_row}')
    
    # ===== Sheet 9: 音游化扩展 =====
    ws9 = wb.create_sheet("9.音游化扩展")
    style_header(ws9, "🎮 音游化扩展设计")
    
    ws9.append(["特色玩法方向"])
    ws9.merge_cells('A3:J3')
    ws9['A3'].font = subtitle_font
    ws9.append([])
    ws9.append(["方向", "玩法说明", "差异化特色"])
    ws9.merge_cells('D5:J5')
    ws9['A5'].font = Font(bold=True)
    ws9['B5'].font = Font(bold=True)
    ws9['C5'].font = Font(bold=True)
    ws9.append(["和声解密", "显示和弦名称，玩家同时按下多个钟的正/侧面拼出和弦", "别的音游没有「一钟双音」设计"])
    ws9.merge_cells('D6:J6')
    ws9.append(["古乐还原", "播放编钟实际演奏录音，玩家跟随演奏", "真实历史乐器「跟弹」体验"])
    ws9.merge_cells('D7:J7')
    ws9.append(["对位竞技", "双人分屏，各控制一半音区合奏", "天然适合「一人旋律、一人伴奏」分工"])
    ws9.merge_cells('D8:J8')
    ws9.append([])
    
    ws9.append(["谱面可视化方案"])
    ws9.merge_cells('A10:J10')
    ws9['A10'].font = subtitle_font
    ws9.append([])
    ws9.append(["方案", "描述"])
    ws9.merge_cells('C12:J12')
    ws9['A12'].font = Font(bold=True)
    ws9['B12'].font = Font(bold=True)
    ws9.append(["分层轨道", "上层=正敲音符(实心圆)，下层=侧敲音符(空心圆)，同时触发=双音提示"])
    ws9.merge_cells('C13:J13')
    ws9.append(["钟体可视化", "谱面即编钟俯视图，音符落入对应钟位置，落点偏左=侧敲，偏右=正敲"])
    ws9.merge_cells('C14:J14')
    ws9.append([])
    
    ws9.append(["难度设计"])
    ws9.merge_cells('A16:J16')
    ws9['A16'].font = subtitle_font
    ws9.append([])
    ws9.append(["难度", "机制", "示例"])
    ws9.merge_cells('D18:J18')
    ws9['A18'].font = Font(bold=True)
    ws9['B18'].font = Font(bold=True)
    ws9['C18'].font = Font(bold=True)
    ws9.append(["Easy", "单音旋律，只用到正敲", "《小星星》"])
    ws9.merge_cells('D19:J19')
    ws9.append(["Normal", "简单和弦，偶尔双音", "三和弦分解"])
    ws9.merge_cells('D20:J20')
    ws9.append(["Hard", "密集双音，跨区快速切换", "巴赫风格对位"])
    ws9.merge_cells('D21:J21')
    ws9.append(["Master", "复节奏，全音区同时触发", "完整交响编配"])
    ws9.merge_cells('D22:J22')
    ws9.append([])
    
    ws9.append(["UGC生态设计（演奏→谱面→分享）"])
    ws9.merge_cells('A24:J24')
    ws9['A24'].font = subtitle_font
    ws9.append([])
    ws9.append(["环节", "说明"])
    ws9.merge_cells('C26:J26')
    ws9['A26'].font = Font(bold=True)
    ws9['B26'].font = Font(bold=True)
    ws9.append(["创作者演奏", "高端玩家自由演奏/即兴创作"])
    ws9.merge_cells('C27:J27')
    ws9.append(["生成谱面", "系统自动量化时间戳、推算BPM、标定难度(NPS计算)"])
    ws9.merge_cells('C28:J28')
    ws9.append(["分享分发", "谱面+预览音频+热力图缩略+通关率统计"])
    ws9.merge_cells('C29:J29')
    ws9.append(["玩家挑战", "普通玩家游玩获得评分(Precision/Good/Miss)"])
    ws9.merge_cells('C30:J30')
    ws9.append(["数据反馈", "通关率、retry分布、点赞数反馈给创作者优化谱面"])
    ws9.merge_cells('C31:J31')
    
    # ===== Sheet 10: 参考资料 =====
    ws10 = wb.create_sheet("10.参考资料")
    style_header(ws10, "📎 附录与参考")
    
    ws10.append(["参考资源"])
    ws10.merge_cells('A3:J3')
    ws10['A3'].font = subtitle_font
    ws10.append([])
    ws10.append(["曾侯乙编钟 - 湖北省博物馆"])
    ws10.merge_cells('A5:J5')
    ws10.append(["Web Audio API规范 - W3C"])
    ws10.merge_cells('A6:J6')
    ws10.append(["《中国古代音乐史稿》杨荫浏"])
    ws10.merge_cells('A7:J7')
    ws10.append([])
    
    ws10.append(["相关链接"])
    ws10.merge_cells('A9:J9')
    ws10['A9'].font = subtitle_font
    ws10.append([])
    ws10.append(["在线体验", "https://hiyascott.github.io/scott-portfolio/research/instrument-simulator/bianzhong/"])
    ws10.merge_cells('B10:J10')
    ws10.append(["设计案文档", "https://hiyascott.github.io/scott-portfolio/research/instrument-simulator/bianzhong/design-doc.html"])
    ws10.merge_cells('B11:J11')
    ws10.append(["能力图谱", "https://hiyascott.github.io/scott-portfolio/kimi-claw/"])
    ws10.merge_cells('B12:J12')
    ws10.append([])
    
    ws10.append(["文档信息"])
    ws10.merge_cells('A14:J14')
    ws10['A14'].font = subtitle_font
    ws10.append([])
    ws10.append(["版本", "v1.0.0"])
    ws10.merge_cells('B15:J15')
    ws10.append(["日期", "2026.03.11"])
    ws10.merge_cells('B16:J16')
    ws10.append(["作者", "Kimi Claw (AI Assistant)"])
    ws10.merge_cells('B17:J17')
    
    # 设置列宽
    for ws in wb.worksheets:
        if ws.title not in ['1.项目概述', '2.核心机制', '3.音频配置表', '4.UI-UE设计']:
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 25
            for col in ['C','D','E','F','G','H','I','J']:
                ws.column_dimensions[col].width = 15
    
    # 保存
    output_path = '/root/.openclaw/workspace/编钟模拟器设计案_完整版.xlsx'
    wb.save(output_path)
    print(f"✅ 完整版Excel已更新: {output_path}")
    print(f"📑 共 {len(wb.worksheets)} 个工作表:")
    for i, ws in enumerate(wb.worksheets, 1):
        print(f"   {i}. {ws.title}")
    return output_path

if __name__ == '__main__':
    add_remaining_sheets()
